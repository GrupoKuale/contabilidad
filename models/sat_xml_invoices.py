# -*- coding: utf-8 -*-
import gzip
import zlib

from odoo import api, fields, models, tools, _
from odoo.exceptions import ValidationError, UserError
from lxml import etree
import base64
import logging
import requests
from unittest.mock import patch
import qrcode

from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.utils import ImageReader
from datetime import datetime
import io

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None

_logger = logging.getLogger(__name__)


class SATXMLInvoicesImpuestos(models.Model):
    _name = 'sat.invoices.taxes'
    _description = "SAT XML Invoices Taxes"

    name = fields.Char('Impuesto', required=True)
    tipoFactor = fields.Char('Tipo Factor')
    sat_concept_id = fields.Many2one('sat.invoices.concepts', string='Concepto SAT')
    base = fields.Float('Base')
    tasa_cuota = fields.Float('Tasa o Cuota')
    importe = fields.Float('Importe')


class SATXMLInvoicesImpuestos(models.Model):
    _name = 'sat.invoices.taxes2'
    _description = "SAT XML Invoices Taxes 2"

    name = fields.Char('Impuesto', required=True)
    tipoFactor = fields.Char('Tipo Factor')
    sat_concept_id = fields.Many2one('sat.invoices.concepts', string='Concepto SAT')
    base = fields.Float('Base')
    tasa_cuota = fields.Float('Tasa o Cuota')
    importe = fields.Float('Importe')


class SATXMLInvoicesConceptos(models.Model):
    _name = 'sat.invoices.concepts'
    _description = "SAT XML Invoices Concepts"

    name = fields.Char('Producto SAT', required=True)
    description = fields.Char('Descripción')
    clave_unidad = fields.Char('Clave Unidad')
    no_id = fields.Char('No. Identificación')
    objeto_imp = fields.Char('Objeto Imp.')
    sat_invoice_id = fields.Many2one('sat.xml.invoices', string='Factura SAT')
    cantidad = fields.Float('Cantidad')
    valor_unitario = fields.Float('Valor Unitario')
    importe = fields.Float('Importe')
    traslados_ids = fields.One2many('sat.invoices.taxes', 'sat_concept_id', string='Traslados')
    retenciones_ids = fields.One2many('sat.invoices.taxes2', 'sat_concept_id', string='Retenciones')

class SATXMLInvoiceComplements(models.Model):
    _name = 'sat.invoices.complements'
    _description = "SAT XML Invoice Complementos"
    name = fields.Char('Nombre del Complemento', required=True)
    tipo_complemento = fields.Selection([
        ('pagos', 'Pagos'),
        ('nomina', 'Nómina'),
        ('carta_porte', 'Carta Porte'),
        ('impuestos_locales', 'Impuestos Locales'),
        ('ine', 'INE'),
        ('notarios_publicos', 'Notarios Públicos'),
        ('donatarias', 'Donatarias'),
        ('leyendas_fiscales', 'Leyendas Fiscales'),
        ('turista_extranjero', 'Turista Extranjero'),
        ('otros', 'Otros')
    ], string="Tipo de Complemento", required=True)

    sat_invoice_id = fields.Many2one('sat.xml.invoices', string='Factura SAT', required=True)
    fecha = fields.Date('Fecha Complemento')
    moneda = fields.Char('Moneda')
    total = fields.Float('Total')

    traslados_ids = fields.One2many('sat.invoices.complement.taxes', 'sat_complement_id', string='Traslados')
    retenciones_ids = fields.One2many('sat.invoices.complement.taxes', 'sat_complement_id', string='Retenciones')

class SATXMLInvoiceComplementTaxes(models.Model):
    _name = 'sat.invoices.complement.taxes'
    _description = "SAT XML Invoice Complement Taxes"

    name = fields.Char('Impuesto', required=True)
    tipoFactor = fields.Char('Tipo Factor')
    base = fields.Float('Base')
    tasa_cuota = fields.Float('Tasa o Cuota')
    importe = fields.Float('Importe')

    tipo = fields.Selection([
        ('traslado', 'Traslado'),
        ('retencion', 'Retención')
    ], string="Tipo de Impuesto", required=True)

    sat_complement_id = fields.Many2one('sat.invoices.complements', string='Complemento SAT', required=True)


class SATXMLInvoices(models.Model):
    _name = 'sat.xml.invoices'
    _description = "SAT XML Invoices"
    _rec_names_search = ['name','tfd_uuid']
    _rec_name = 'tfd_uuid'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    
    active = fields.Boolean('Activo', default=True, tracking=True,
                           help='Si está desactivado, el documento está archivado y no aparece en vistas normales')
    
    name = fields.Char('Referencia', required=True)
    tfd_uuid = fields.Char('UUID', index=True)
    rfc_emisor = fields.Char('RFC Emisor', index=True)
    rfc_receptor = fields.Char('RFC Receptor', index=True)
    xml_file_original = fields.Binary(string='XML Original',readonly=True)
    xml_file = fields.Binary(string='XML')
    xml_file_name = fields.Char('XML')
    nombre_emisor = fields.Char('Nombre Emisor')
    nombre_receptor = fields.Char('Nombre Receptor')
    factura_version = fields.Char('Versión')
    factura_serie = fields.Char('Serie')
    factura_folio = fields.Char('Folio')
    factura_fecha = fields.Date('Fecha Comprobante', index=True)
    factura_forma_pago = fields.Char('Forma de Pago')
    factura_subtotal = fields.Float('Subtotal')
    factura_impuestos = fields.Float('Impuestos')
    factura_total = fields.Float('Total')
    factura_tipo_cambio = fields.Float('Tipo de Cambio', default=1.00)
    factura_moneda = fields.Char('Moneda')
    factura_tipo = fields.Char('Tipo de Comprobante')
    factura_exportacion = fields.Char('Exportación')
    factura_metodo_pago = fields.Char('Método Pago')
    factura_lugar_expedicion = fields.Char('Lugar Expedición')
    
    sat_cuenta = fields.Many2one('account.account', string='Cuenta Contable')
    sat_cuenta1 = fields.Many2one('account.account', string='Cuenta Contable base')
    sat_cuenta2 = fields.Many2one('account.account', string='Cuenta Contable sub1')
    sat_cuenta3 = fields.Many2one('account.account', string='Cuenta Contable sub2')
    partner_id = fields.Many2one('res.partner', string='Proveedor')

    referencia = fields.Char(
        string='Referencia',
        tracking=True,
        help='Campo abierto para que el contador escriba referencias'
    )

    observaciones = fields.Text(
        string='Observaciones',
        tracking=True,
        help='Campo abierto para observaciones'
    )

    validez = fields.Char(
        string='Validez',
        help='Estatus de validación del XML'
    )

    
    concepts_ids = fields.One2many('sat.invoices.concepts', 'sat_invoice_id', string='Conceptos')
    complements_ids = fields.One2many('sat.invoices.complements', 'sat_invoice_id', string='Complementos')
    credit_note_ids = fields.One2many(
        'account.move','invoice',
        string='Notas de Crédito / Bonificaciones',
        help='Notas de crédito o bonificaciones asociadas a esta factura'
    )
    
    state_detail = fields.Selection([
        ('on_wait', 'En espera'),
        ('done', 'Procesada'),
        ('error', 'Error'),
    ], string='Estatus de importación', default='on_wait')

    is_external_document = fields.Boolean(string='Documento Externo', default=False,  help='Indica que este registro no proviene de un CFDI XML')

    # Campo para identificar si es emitido o recibido
    document_type = fields.Selection([
        ('recibido', 'XML Recibido'),
        ('emitido', 'XML Emitido'),
        ('documento_externo', 'Documento'),
    ], string='Tipo de Documento', required=True, default='recibido', index=True,
       help='Indica si el CFDI fue emitido por la empresa o recibido de un proveedor')
    
    # Clasificación del documento (Factura, Pago, Nómina, etc.)
    document_group = fields.Selection([
        ('factura', 'Factura'),
        ('pago', 'Pago'),
        ('nomina', 'Nómina'),
        ('retencion', 'Retención'),
        ('traslado', 'Traslado'),
        ('egreso', 'Egreso'),
    ], string='Grupo de Documento', compute='_compute_document_group', store=True,
       help='Clasificación del CFDI según su tipo de comprobante')
    
    # Campos de control ADD
    add_user_id = fields.Many2one(
        'res.users', 
        string='Usuario ADD',
        index=True,
        help='Usuario que tiene este documento en su ADD',
        ondelete='set null'
    )
    
    add_status = fields.Selection([
        ('available', 'Disponible'),
        ('locked', 'Bloqueado en ADD'),
        ('processed', 'Procesado'),
    ], string='Estado ADD', default='available', index=True,
       help='Estado del documento en el sistema ADD')
    
    add_date = fields.Datetime(
        string='Fecha Carga ADD',
        help='Fecha en que se cargó el documento al ADD'
    )
    
    # Verificación ante el SAT
    sat_status = fields.Selection([
        ('vigente', 'Vigente'),
        ('cancelado', 'Cancelado'),
        ('no_verificado', 'No Verificado'),
    ], string='Estado SAT', default='no_verificado', index=True,
       help='Estado del CFDI ante el SAT')
    
    sat_verification_date = fields.Datetime(
        string='Última Verificación SAT',
        help='Fecha de la última verificación de estado ante el SAT'
    )
    
    sat_cancellation_date = fields.Datetime(
        string='Fecha de Cancelación',
        help='Fecha en que el CFDI fue cancelado ante el SAT'
    )

    # Verificación de estatus para el sistema PK
    pk_status_received = fields.Selection([
        ('sin_asignacion','Sin Asignacion'),
        ('asignada','Asignada'),
        ('vinculacion_pendiente', 'Vinculacion Pendiente'),
        ('sin_oc', 'Sin OC'),
        ('oc_por_autorizar', 'OC Por Autorizar'),
        ('pago_pendiente', 'Pago Pendiente'),
        ('pagada_terminada', 'Pagada Terminada'),
        ('pagada_REP_pendiente', 'Pagada REP Pendiente'),
        ('pagada_REP_terminado', 'Pagada REP Terminado'),
    ], string="Estado PK ", default='sin_asignacion',
        index=True, help="Estado de la Factura Recibida Contable para Punto Kuale")

    pk_status_sent = fields.Selection([
        ('por_cobrar','Por Cobrar'),
        ('cobrada_terminado','Cobrada Terminado'),
        ('cobrada_rep_pendiente', 'Cobrada REP Pendiente'),
        ('cobrada_rep_terminado', 'Cobrada REP Terminado'),
    ], string="Estado PK ", default='por_cobrar',
        index=True, help="Estado de la Factura Emitada Contable para Punto Kuale")

    # Campo computado para saber si está disponible
    is_available = fields.Boolean(
        string='Disponible para ADD',
        compute='_compute_is_available',
        search='_search_is_available',
        help='Indica si el documento está disponible para ser agregado a un ADD'
    )

    
    @api.depends('factura_tipo', 'complements_ids.tipo_complemento')
    def _compute_document_group(self):
        """Clasifica el documento según su tipo de comprobante"""
        for record in self:
            tipo = record.factura_tipo
            
            if tipo == 'I':  # Ingreso
                record.document_group = 'factura'
            elif tipo == 'E':  # Egreso
                record.document_group = 'egreso'
            elif tipo == 'P':  # Pago
                record.document_group = 'pago'
            elif tipo == 'N':  # Nómina
                record.document_group = 'nomina'
            elif tipo == 'T':  # Traslado
                record.document_group = 'traslado'
            else:
                # Verificar si tiene complemento de pago
                if record.complements_ids.filtered(lambda c: c.tipo_complemento == 'pagos'):
                    record.document_group = 'pago'
                elif record.complements_ids.filtered(lambda c: c.tipo_complemento == 'nomina'):
                    record.document_group = 'nomina'
                else:
                    record.document_group = 'factura'
    
    @api.depends('add_status', 'add_user_id')
    def _compute_is_available(self):
        """Determina si el documento está disponible para ADD"""
        for record in self:
            record.is_available = (
                record.add_status == 'available' and 
                not record.add_user_id
            )
    
    def _search_is_available(self, operator, value):
        """Búsqueda personalizada para documentos disponibles"""
        if operator == '=' and value:
            return [('add_status', '=', 'available'), ('add_user_id', '=', False)]
        elif operator == '=' and not value:
            return ['|', ('add_status', '!=', 'available'), ('add_user_id', '!=', False)]
        return []

    @api.constrains('is_external_document','rfc_emisor','nombre_emisor','rfc_receptor','nombre_receptor')
    def _check_required_fields_for_cfdi(self):
        for rec in self:
            if not rec.is_external_document:
                missing = []
                if not rec.rfc_emisor:
                    missing.append('RFC Emisor')
                if not rec.nombre_emisor:
                    missing.append('Nombre Emisor')
                if not rec.rfc_receptor:
                    missing.append('RFC Receptor')
                if not rec.nombre_receptor:
                    missing.append('Nombre Receptor')

                if missing:
                    raise ValidationError(
                        'Faltan campos obligatorios para CFDI:\n- ' +
                        '\n- '.join(missing)
                    )

    _sql_constraints = [
        ('uuid_unique', 'UNIQUE(tfd_uuid)', 
         'El UUID ya existe en el sistema. No se pueden duplicar facturas.'),
    ]

    
    @api.constrains('add_user_id', 'add_status')
    def _check_add_user_lock(self):
        """Valida que solo un usuario pueda tener el documento en su ADD"""
        for record in self:
            if record.add_user_id and record.add_status != 'available':
                # Verificar si hay otro registro con el mismo UUID pero diferente usuario
                conflicting = self.search([
                    ('id', '!=', record.id),
                    ('tfd_uuid', '=', record.tfd_uuid),
                    ('add_user_id', '!=', False),
                    ('add_user_id', '!=', record.add_user_id.id),
                    ('add_status', 'in', ['locked', 'processed'])
                ])
                if conflicting:
                    raise ValidationError(_(
                        'El documento con UUID %s ya está siendo utilizado por %s en su ADD.'
                    ) % (record.tfd_uuid, conflicting[0].add_user_id.name))

    @api.model
    def create(self, values):
        """Método create modificado para parsear XML y detectar tipo de documento"""
        filename = values.get('xml_file_name') or ''
        xml_b64 = values.get('xml_file')
        
        if filename.endswith('.xml') and xml_b64:
            try:
                file_data = base64.b64decode(xml_b64)
                root = etree.fromstring(file_data)
                ns = {
                    'cfdi': 'http://www.sat.gob.mx/cfd/4',
                    'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital',
                }
                
                # -------- Comprobante --------
                comp = root
                values['factura_version'] = comp.get('Version')
                values['factura_serie'] = comp.get('Serie')
                values['factura_folio'] = comp.get('Folio')
                values['factura_fecha'] = comp.get('Fecha')
                values['factura_forma_pago'] = comp.get('FormaPago')
                values['factura_moneda'] = comp.get('Moneda')
                values['factura_tipo'] = comp.get('TipoDeComprobante')
                values['factura_exportacion'] = comp.get('Exportacion')
                values['factura_metodo_pago'] = comp.get('MetodoPago')
                values['factura_lugar_expedicion'] = comp.get('LugarExpedicion')
                
                subtotal = float(comp.get('SubTotal') or '0')
                descuento = float(comp.get('Descuento') or '0')
                total = float(comp.get('Total') or '0')
                
                values['factura_subtotal'] = subtotal - descuento
                values['factura_total'] = total
                values['factura_impuestos'] = total - values['factura_subtotal']
                values['factura_tipo_cambio'] = float(comp.get('TipoCambio') or '1')
                
                # -------- Emisor / Receptor --------
                emisor = root.find('cfdi:Emisor', ns)
                receptor = root.find('cfdi:Receptor', ns)
                
                values.setdefault('rfc_emisor', 'DESCONOCIDO')
                values.setdefault('nombre_emisor', 'DESCONOCIDO')
                values.setdefault('rfc_receptor', 'DESCONOCIDO')
                values.setdefault('nombre_receptor', 'DESCONOCIDO')
                
                if emisor is not None:
                    if emisor.get('Rfc'):
                        values['rfc_emisor'] = emisor.get('Rfc')
                    if emisor.get('Nombre'):
                        values['nombre_emisor'] = emisor.get('Nombre')
                
                if receptor is not None:
                    if receptor.get('Rfc'):
                        values['rfc_receptor'] = receptor.get('Rfc')
                    if receptor.get('Nombre'):
                        values['nombre_receptor'] = receptor.get('Nombre')
                
                # -------- DETECTAR TIPO DE DOCUMENTO --------
                company_rfc = self.env.company.rfc
                if company_rfc:
                    # Si la empresa es el emisor, es un documento emitido
                    if values['rfc_emisor'] == company_rfc:
                        values['document_type'] = 'emitido'
                    # Si la empresa es el receptor, es un documento recibido
                    elif values['rfc_receptor'] == company_rfc:
                        values['document_type'] = 'recibido'
                
                # -------- UUID (timbre) --------
                tfd = root.xpath('.//tfd:TimbreFiscalDigital', namespaces=ns)
                if tfd:
                    values['tfd_uuid'] = tfd[0].get('UUID')
                
                # Validar que exista UUID antes de crear
                if not values.get('tfd_uuid'):
                    _logger.error("XML %s sin TimbreFiscalDigital/UUID; se omite creación.", filename)
                    return self.env['sat.xml.invoices']
                
                # -------- Chequeo de duplicados --------
                dup = self.env['sat.xml.invoices'].search([('tfd_uuid', '=', values['tfd_uuid'])], limit=1)
                if dup:
                    # Actualizar solo si no está bloqueado en un ADD
                    if dup.add_status == 'locked' and dup.add_user_id:
                        _logger.warning(
                            "Factura UUID=%s bloqueada por usuario %s. No se actualiza.",
                            values['tfd_uuid'], dup.add_user_id.name
                        )
                        return dup
                    
                    dup.write(values)
                    _logger.info("Factura SAT existente actualizada (UUID=%s, id=%s)", values['tfd_uuid'], dup.id)
                    return dup
                
                # -------- Conceptos / Impuestos --------
                concepts_vals = []
                conceptos = root.xpath('cfdi:Conceptos/cfdi:Concepto', namespaces=ns)
                
                for c in conceptos:
                    traslados_vals = []
                    retenciones_vals = []
                    
                    impuestos = c.find('cfdi:Impuestos', ns)
                    if impuestos is not None:
                        traslados = impuestos.xpath('cfdi:Traslados/cfdi:Traslado', namespaces=ns)
                        for t in traslados:
                            traslados_vals.append((0, 0, {
                                'name': t.get('Impuesto'),
                                'tipoFactor': t.get('TipoFactor'),
                                'base': t.get('Base'),
                                'tasa_cuota': t.get('TasaOCuota'),
                                'importe': t.get('Importe'),
                            }))
                        
                        retenciones = impuestos.xpath('cfdi:Retenciones/cfdi:Retencion', namespaces=ns)
                        for r in retenciones:
                            retenciones_vals.append((0, 0, {
                                'name': r.get('Impuesto'),
                                'tipoFactor': r.get('TipoFactor'),
                                'base': r.get('Base'),
                                'tasa_cuota': r.get('TasaOCuota'),
                                'importe': r.get('Importe'),
                            }))
                    
                    concepts_vals.append((0, 0, {
                        'name': c.get('ClaveProdServ'),
                        'description': c.get('Descripcion'),
                        'clave_unidad': c.get('ClaveUnidad'),
                        'no_id': c.get('NoIdentificacion'),
                        'objeto_imp': c.get('ObjetoImp'),
                        'cantidad': c.get('Cantidad'),
                        'valor_unitario': c.get('ValorUnitario'),
                        'importe': c.get('Importe'),
                        'traslados_ids': traslados_vals,
                        'retenciones_ids': retenciones_vals,
                    }))
                
                values['concepts_ids'] = concepts_vals
                
            except Exception as e:
                _logger.error("Error parseando XML %s: %s", filename, e)

        # ----- Obtencion del XML para poder Mostrarlo -----
        values['xml_file_original'] = values.get('xml_file')
        # ---------------- CREACIÓN DEL REGISTRO ----------------
        # -------- CREACIÓN DEL REGISTRO ----------------
        invoice = super(SATXMLInvoices, self).create(values)

    # 1. Si es XML RECIBIDO (Compras) → intentamos automatizar la OC
        if invoice.document_type == 'recibido':
            if not self.env.context.get('skip_oc_creation'):
                try:
                    wizard = self.env['purchase.order.import.xml'].sudo().create({
                        'file': invoice.xmlfile,
                        'filename': invoice.xmlfilename,
                    })
                    
                    result = wizard.sudo().action_process_massive_xml()
                    
                    if isinstance(result, dict) and result.get('type') == 'ir.actions.act_window':
                        # Es un wizard de productos faltantes → onwait
                        invoice.state_detail = 'onwait'
                        _logger.info(f" XML {invoice.xmlfilename} requiere conciliación manual")
                        
                    elif result:  # Es una orden de compra (objeto)
                        invoice.state_detail = 'done'
                        _logger.info(f"OC {result.name} creada automáticamente para UUID {invoice.tfduuid}")
                        
                        if result.requires_authorization:
                            try:
                                result.action_request_approval()
                                _logger.info(f" Flujo de aprobación iniciado automáticamente para {result.name}")
                            except Exception as e:
                                _logger.error(f" Error al solicitar aprobación automática: {e}")
                    else:
                        invoice.state_detail = 'onwait'
                        
                except Exception as e:
                    invoice.state_detail = 'error'
                    _logger.error(f" Error creando OC automática para {invoice.xmlfilename}: {e}")
            else:
                invoice.state_detail = 'done'

        elif invoice.document_type == 'emitido':
            invoice.state_detail = 'onwait'  #
        else:
            invoice.state_detail = 'done'


        # Validación SAT 
        if not self.env.context.get('skip_sat_verify') and invoice.tfd_uuid and invoice.sat_status == 'no_verificado':
            try:
                invoice.action_verify_sat_status()
            except UserError as e:
                _logger.warning("Verificación SAT omitida al crear factura UUID=%s: %s", invoice.tfd_uuid, e)
            except Exception as e:
                _logger.error("Error inesperado verificando SAT al crear factura UUID=%s: %s", invoice.tfd_uuid, e)

        return invoice

    def _get_sello_from_xml(self, record):
        """
        Extrae el sello digital del XML (cfdi:Comprobante)
        """
        if not record.xml_file:
            raise UserError(_("La factura no tiene XML cargado."))

        xml_data = base64.b64decode(record.xml_file)
        root = etree.fromstring(xml_data)

        sello = root.get("Sello")
        if not sello:
            raise UserError(_("No se encontró el atributo Sello en el XML."))

        return sello

    def action_add_to_user_add(self, user_id=None):
        """
        Agrega el documento al ADD del usuario
        """
        self.ensure_one()
        
        if not user_id:
            user_id = self.env.user.id
        
        if self.add_status == 'locked' and self.add_user_id:
            if self.add_user_id.id != user_id:
                raise UserError(_(
                    'Este documento ya está siendo utilizado por %s en su ADD. '
                    'No puede agregarlo hasta que sea liberado.'
                ) % self.add_user_id.name)
            else:
                raise UserError(_('Este documento ya está en su ADD.'))
        
        self.write({
            'add_user_id': user_id,
            'add_status': 'locked',
            'add_date': fields.Datetime.now(),
        })
        
        _logger.info('Documento %s agregado al ADD de usuario ID=%s', self.tfd_uuid, user_id)
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Documento Agregado'),
                'message': _('El documento ha sido agregado a su ADD correctamente.'),
                'type': 'success',
                'sticky': False,
            }
        }
    
    def action_remove_from_add(self):
        """
        Remueve el documento del ADD del usuario
        """
        self.ensure_one()
        
        if not self.add_user_id or self.add_user_id.id != self.env.user.id:
            raise UserError(_('Solo el usuario propietario puede remover este documento de su ADD.'))
        
        self.write({
            'add_user_id': False,
            'add_status': 'available',
            'add_date': False,
        })
        
        _logger.info('Documento %s removido del ADD', self.tfd_uuid)
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Documento Liberado'),
                'message': _('El documento ha sido removido de su ADD y está disponible nuevamente.'),
                'type': 'info',
                'sticky': False,
            }
        }

    def action_verify_sat_status(self):
        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        API_URL = f"{base_url}/api/cfdi/validate"
        config = self.env['ir.config_parameter'].sudo()
        sat_available = self.env['ir.config_parameter'].sudo().get_param('sat.service.available', 'true')
        last_check = config.get_param('sat.service.last_check')

        if sat_available == 'false' and last_check:
            last_check_dt = fields.Datetime.from_string(last_check)
            if (fields.Datetime.now() - last_check_dt).total_seconds() < 300:
                raise UserError(_(
                    "El sistema de validación del SAT no está disponible.\n"
                    "Por favor intente nuevamente más tarde."
                ))

        for record in self:

            if not record.tfd_uuid:
                raise UserError(_("La factura no tiene UUID para validar ante el SAT."))

            payload = {
                "uuid": record.tfd_uuid,
                "rfc_emisor": record.rfc_emisor,
                "rfc_receptor": record.rfc_receptor,
                "total": record.factura_total,
                "sello": self._get_sello_from_xml(record),
            }

            try:
                response = requests.post(API_URL, json=payload, timeout=5)

                if response.status_code in (500, 502, 503, 504):
                    raise requests.exceptions.RequestException(f"HTTP {response.status_code}")

                response.raise_for_status()
                data = response.json()

            except Exception as e:

                self._mark_sat_service_down(str(e))

                record.write({'sat_verification_date': fields.Datetime.now(),})

                raise UserError(_(
                    "El sistema de validación del SAT no está disponible en este momento.\n"
                    "No se pudieron verificar las facturas.\n\n"
                    "Intente nuevamente más tarde."
                ))

            # Hay respuesta del SAT
            self._mark_sat_service_up()

            if data.get('success'):
                estado_sat = (data.get('estado') or '').strip().lower()

                status_map = {
                    'vigente': 'vigente',
                    'cancelado': 'cancelado',
                }

                record.write({
                    'sat_status': status_map.get(estado_sat, 'no_verificado'),
                    'sat_verification_date': fields.Datetime.now(),
                    'sat_cancellation_date': (
                        fields.Datetime.now() if estado_sat == 'cancelado' else False
                    )
                })

            else:
                record.write({
                    'sat_status': 'no_verificado',
                    'sat_verification_date': fields.Datetime.now(),
                })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'sat.xml.invoices',
            'view_mode': 'tree,form',
            'name': 'Facturas SAT XML',
            'target': 'current',
        }

    def action_retry_create_oc(self):
        """
        Reintentar creación de OC.
        Si devuelve un Wizard (diccionario), detiene el ciclo y abre la ventana.
        Si devuelve una OC (objeto), continúa con el siguiente.
        """
        invoices_to_retry = self.filtered(lambda x: x.state_detail in ('error', 'on_wait'))
        
        if not invoices_to_retry:
            raise UserError(_("No hay facturas en estado 'Error' o 'En Espera' para reintentar."))
        
        success_count = 0
        error_count = 0
        skipped_count = 0
        
        wizard_action = False

        for invoice in invoices_to_retry:

            if wizard_action:
                break

            savepoint = f"sp_{invoice.id}_{int(fields.Datetime.now().timestamp())}"
            
            try:
                self.env.cr.execute(f"SAVEPOINT {savepoint}")
                
                # Validar que tenga XML
                if not invoice.xml_file:
                    _logger.warning(f"Factura {invoice.tfd_uuid} sin archivo XML, omitiendo...")
                    skipped_count += 1
                    self.env.cr.execute(f"RELEASE SAVEPOINT {savepoint}")
                    continue
                
                # Crear wizard de importación
                wizard = self.env['purchase.order.import.xml'].sudo().create({
                    'file': invoice.xml_file,
                    'filename': invoice.xml_file_name or f"{invoice.tfd_uuid}.xml",
                })
                
                # Ejecutar proceso
                result = wizard.sudo().action_process_massive_xml()
                
                # --- WIZARD DE CONCILIACIÓN ---
                if isinstance(result, dict) and result.get('type') == 'ir.actions.act_window':
                    self.env.cr.execute(f"RELEASE SAVEPOINT {savepoint}")
                    wizard_action = result
                    break

                # --- ORDEN DE COMPRA CREADA ---
                elif result: 
                    invoice.write({'state_detail': 'done'})
                    success_count += 1
                    _logger.info(f"OC {result.name} creada para UUID {invoice.tfd_uuid}")
                    self.env.cr.execute(f"RELEASE SAVEPOINT {savepoint}")

                # --- FALLÓ ---
                else:
                    invoice.write({'state_detail': 'on_wait'})
                    skipped_count += 1
                    _logger.warning(f"No se generó OC para UUID {invoice.tfd_uuid}")
                    self.env.cr.execute(f"RELEASE SAVEPOINT {savepoint}")
                
            except Exception as e:
                # Rollback al savepoint
                try:
                    self.env.cr.execute(f"ROLLBACK TO SAVEPOINT {savepoint}")
                except Exception as rollback_error:
                    _logger.error(f"Error haciendo rollback: {rollback_error}")
                    self.env.cr.rollback() 
                
                # Marcar como error
                try:
                    invoice.write({'state_detail': 'error'})
                except:
                    pass
                
                error_count += 1
                _logger.error(f"Error reintentando OC (UUID: {invoice.tfd_uuid}): {str(e)}", exc_info=True)
        
        # Commit final de lo que sí se procesó
        try:
            self.env.cr.commit()
        except Exception as e:
            self.env.cr.rollback()
        
        # Si se detectó que se necesita abrir un wizard, lo retornamos ahora
        if wizard_action:
            return wizard_action

        # Si no hubo wizard, mostramos el resumen
        message = _(
            f"Proceso completado:\n"
            f"Éxitos: {success_count}\n"
            f"Errores: {error_count}\n"
            f"Omitidos: {skipped_count}"
        )
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Reintento de Creación de OC'),
                'message': message,
                'type': 'success' if error_count == 0 else ('warning' if success_count > 0 else 'danger'),
                'sticky': True,
            }
        }

    def action_process_document(self):
        """
        Método inteligente: Decide si crea Compra o Venta según el tipo de documento.
        """
        self.ensure_one()
        if not self.xml_file:
             raise UserError(_("No hay archivo XML adjunto."))

        # Validar tipo de documento
        if not self.document_type:
             raise UserError("No se ha identificado si el documento es Emitido o Recibido.")

        # --- XML RECIBIDO (Compras) ---
        if self.document_type == 'recibido':
            _logger.info(f"Procesando XML {self.tfd_uuid} como COMPRA")
            
            wizard = self.env['purchase.order.import.xml'].create({
                'file': self.xml_file,
                'filename': self.xml_file_name or 'factura.xml'
            })
            
            res = wizard.action_process_massive_xml()
            
            if isinstance(res, dict):
                return res
            
            if res:
                self.write({'state_detail': 'done'})
                return True

        # --- XML EMITIDO (Ventas) ---
        elif self.document_type == 'emitido':
            _logger.info(f"Procesando XML {self.tfd_uuid} como VENTA")
            
            wizard = self.env['sale.order.import.xml'].create({
                'file': self.xml_file,
                'filename': self.xml_file_name or 'factura_venta.xml'
            })
            
            sale_order = wizard.action_process_sale_xml()
            
            if sale_order:
                self.write({'state_detail': 'done'})
                
                return {
                    'name': _('Pedido de Venta Creado'),
                    'type': 'ir.actions.act_window',
                    'res_model': 'sale.order',
                    'res_id': sale_order.id,
                    'view_mode': 'form',
                    'target': 'current',
                }

        return False

    def _cron_verify_sat_status(self, batch_size=50):
        """
        Verificación automática de CFDI ante el SAT
        Procesa TODOS los registros en lotes dinámicos
        """
        domain = [
            ('tfd_uuid', '!=', False),
            ('sat_status', 'in', ['no_verificado', 'vigente']),
        ]

        total = self.search_count(domain)

        offset = 0

        while offset < total:
            invoices = self.search(
                domain,
                limit=batch_size,
                offset=offset
            )

            if not invoices:
                break

            # _logger.info(
            #     "Procesando CFDI %s a %s",
            #     offset + 1,
            #     offset + len(invoices)
            # )

            invoices.action_verify_sat_status()

            offset += batch_size

    def _mark_sat_service_down(self, error_msg):
        self.env['ir.config_parameter'].sudo().set_param(
            'sat.service.available', 'false'
        )
        self.env['ir.config_parameter'].sudo().set_param(
            'sat.service.last_error', error_msg
        )
        self.env['ir.config_parameter'].sudo().set_param(
            'sat.service.last_check', fields.Datetime.now()
        )

    def _mark_sat_service_up(self):
        self.env['ir.config_parameter'].sudo().set_param(
            'sat.service.available', 'true'
        )

    def action_verify_sat_status_with_notification(self):
        """
        - Si hay registros seleccionados → valida solo esos
        - Si no hay selección → valida todos
        """
        if self:
            records = self
        else:
            records = self.search([])

        records.action_verify_sat_status()

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Validación completada',
                'message': f'Se validaron {len(records)} factura(s) correctamente.',
                'type': 'success',
                'sticky': False,
            }
        }

    #Mirar XML en codigo
    def action_xml_view_code(self):
        self.ensure_one()

        if not self.xml_file_original:
            raise UserError(
                "Este registro no tiene XML original almacenado.\n"
                "Probablemente fue creado antes de la corrección."
            )

        data = base64.b64decode(self.xml_file_original)
        xml_string = data.decode('utf-8', errors='ignore')

        try:
            etree.fromstring(xml_string.encode('utf-8'))
        except Exception:
            raise UserError("El contenido no es un XML válido.")

        return {
            'type': 'ir.actions.act_window',
            'name': 'Código XML',
            'res_model': 'sat.xml.view.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_xml_code': xml_string,
            }
        }

    @api.model
    def get_available_for_add(self, domain=None):
        """
        Obtiene documentos disponibles para agregar al ADD
        Usado por la vista ContPaq
        """
        base_domain = [
            ('add_status', '=', 'available'),
            ('add_user_id', '=', False)
        ]
        
        if domain:
            base_domain.extend(domain)
        
        return self.search(base_domain)

    @api.model
    def get_my_add_documents(self):
        """
        Obtiene documentos del ADD del usuario actual
        Usado por la vista ContPaq
        """
        return self.search([
            ('add_user_id', '=', self.env.user.id),
            ('add_status', 'in', ['locked', 'processed'])
        ])

    def action_toggle_add(self):
        """
        Alterna el documento entre ADD y disponible
        Usado desde la vista ContPaq
        """
        for record in self:
            if record.add_user_id and record.add_user_id.id == self.env.user.id:
                # Remover del ADD
                record.action_remove_from_add()
            elif record.add_status == 'available' and not record.add_user_id:
                # Agregar al ADD
                record.action_add_to_user_add()
            else:
                raise UserError(_(
                    'El documento está bloqueado por otro usuario (%s)'
                ) % record.add_user_id.name)
            
    def action_generate_xml_pdf(self):
        """
        Genera un PDF del CFDI con el formato estándar de factura ContPaq.
        Incluye todas las secciones: Comprobante, Emisor, Receptor, Conceptos, Totales e Información Fiscal.
        """
        self.ensure_one()
        
        if not self.xml_file:
            raise UserError(_('El documento no tiene archivo XML adjunto'))
        
        try:
            # ========== CONFIGURACIÓN DEL DOCUMENTO ==========
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer, 
                pagesize=letter,
                rightMargin=0.3*inch, 
                leftMargin=0.3*inch,
                topMargin=0.3*inch, 
                bottomMargin=0.5*inch
            )
            
            # ========== DEFINICIÓN DE ESTILOS ==========
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'ComprobanteTitulo',
                parent=styles['Normal'],
                fontSize=18,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#003D82'),
                alignment=TA_CENTER,
                spaceAfter=6
            )
            
            subtitle_style = ParagraphStyle(
                'Subtitulo',
                parent=styles['Normal'],
                fontSize=10,
                fontName='Helvetica',
                textColor=colors.black,
                alignment=TA_CENTER,
                spaceAfter=8
            )
            
            info_label_style = ParagraphStyle(
                'InfoLabel',
                parent=styles['Normal'],
                fontSize=9,
                fontName='Helvetica-Bold',
                textColor=colors.black,
            )
            
            info_value_style = ParagraphStyle(
                'InfoValue',
                parent=styles['Normal'],
                fontSize=9,
                fontName='Helvetica',
                textColor=colors.black,
            )
            
            section_header_style = ParagraphStyle(
                'SectionHeader',
                parent=styles['Normal'],
                fontSize=11,
                fontName='Helvetica-Bold',
                textColor=colors.white,
                backgroundColor=colors.HexColor('#003D82'),
                spaceAfter=6,
                leftIndent=4,
                rightIndent=4,
                leading=14,
            )
            
            story = []
            
            # ========== ENCABEZADO ==========
            story.append(Paragraph("COMPROBANTE FISCAL DIGITAL POR INTERNET (CFDI)", title_style))
            story.append(Spacer(1, 0.15*inch))
            story.append(Paragraph("Representación impresa de un CFDI", subtitle_style))
            story.append(Spacer(1, 0.15*inch))
            
            # ========== INFORMACIÓN DEL COMPROBANTE ==========
            comprobante_data = [
                [Paragraph("<b>Tipo de Comprobante:</b>", info_label_style), 
                Paragraph(f"{self.factura_tipo or 'N/A'} - Ingreso", info_value_style)],
                [Paragraph("<b>Serie:</b>", info_label_style), 
                Paragraph(str(self.factura_serie or 'N/A'), info_value_style)],
                [Paragraph("<b>Folio:</b>", info_label_style), 
                Paragraph(str(self.factura_folio or 'N/A'), info_value_style)],
                [Paragraph("<b>Fecha:</b>", info_label_style), 
                Paragraph(str(self.factura_fecha or 'N/A'), info_value_style)],
                [Paragraph("<b>Lugar de expedición:</b>", info_label_style), 
                Paragraph(str(self.factura_lugar_expedicion or 'N/A'), info_value_style)],
                [Paragraph("<b>Método de Pago:</b>", info_label_style), 
                Paragraph(str(self.factura_metodo_pago or 'N/A'), info_value_style)],
                [Paragraph("<b>Forma de Pago:</b>", info_label_style), 
                Paragraph(str(self.factura_forma_pago or 'N/A'), info_value_style)],
                [Paragraph("<b>Moneda:</b>", info_label_style), 
                Paragraph(f"{self.factura_moneda or 'MXN'}", info_value_style)],
                [Paragraph("<b>Tipo de cambio:</b>", info_label_style), 
                Paragraph(f"{self.factura_tipo_cambio:.4f}", info_value_style)],
            ]
            
            comprobante_table = Table(comprobante_data, colWidths=[2.2*inch, 2.8*inch])
            comprobante_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            
            story.append(comprobante_table)
            story.append(Spacer(1, 0.15*inch))
            
            # ========== EMISOR ==========
            story.append(Paragraph("EMISOR", section_header_style))
            
            emisor_data = [
                [Paragraph("<b>RFC:</b>", info_label_style), 
                Paragraph(str(self.rfc_emisor or 'N/A'), info_value_style)],
                [Paragraph("<b>Nombre:</b>", info_label_style), 
                Paragraph(str(self.nombre_emisor or 'N/A'), info_value_style)],
                [Paragraph("<b>Régimen Fiscal:</b>", info_label_style), 
                Paragraph("601 - General de Ley Personas Morales", info_value_style)],
            ]
            
            emisor_table = Table(emisor_data, colWidths=[2.2*inch, 2.8*inch])
            emisor_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F0F0F0')),
            ]))
            
            story.append(emisor_table)
            story.append(Spacer(1, 0.1*inch))
            
            # ========== RECEPTOR ==========
            story.append(Paragraph("RECEPTOR", section_header_style))
            
            receptor_data = [
                [Paragraph("<b>RFC:</b>", info_label_style), 
                Paragraph(str(self.rfc_receptor or 'N/A'), info_value_style)],
                [Paragraph("<b>Nombre:</b>", info_label_style), 
                Paragraph(str(self.nombre_receptor or 'N/A'), info_value_style)],
                [Paragraph("<b>Uso CFDI:</b>", info_label_style), 
                Paragraph("G01 - Adquisición de mercancías", info_value_style)],
                [Paragraph("<b>Domicilio Fiscal:</b>", info_label_style), 
                Paragraph(str(self.factura_lugar_expedicion or 'N/A'), info_value_style)],
            ]
            
            receptor_table = Table(receptor_data, colWidths=[2.2*inch, 2.8*inch])
            receptor_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F0F0F0')),
            ]))
            
            story.append(receptor_table)
            story.append(Spacer(1, 0.15*inch))
            
            # ========== CONCEPTOS (Tabla de items) ==========
            story.append(Paragraph("CONCEPTOS", section_header_style))
            
            conceptos_headers = [
                'Cantidad',
                'Unidad',
                'Clave',
                'Descripción',
                'Valor Unitario',
                'Descuento',
                'Impuestos',
                'Importe'
            ]
            
            conceptos_data = [
                [Paragraph(f"<b>{h}</b>", ParagraphStyle('h', parent=styles['Normal'], 
                            fontSize=8, fontName='Helvetica-Bold')) for h in conceptos_headers]
            ]
            
            # Agregar conceptos desde la relación one2many
            for concepto in self.concepts_ids:
                conceptos_data.append([
                    Paragraph(str(concepto.cantidad or '0'), info_value_style),
                    Paragraph(str(concepto.clave_unidad or ''), info_value_style),
                    Paragraph(str(concepto.name[:10] or ''), info_value_style),
                    Paragraph(str(concepto.description or '')[:30], info_value_style),
                    Paragraph(f"${concepto.valor_unitario:.2f}" if concepto.valor_unitario else "$0.00", info_value_style),
                    Paragraph("$0.00", info_value_style),
                    Paragraph(f"${sum([t.importe for t in concepto.traslados_ids] or [0]):.2f}", info_value_style),
                    Paragraph(f"${concepto.importe or 0:.2f}", info_value_style),
                ])
            
            conceptos_table = Table(conceptos_data, 
                                colWidths=[0.7*inch, 0.6*inch, 0.5*inch, 1.8*inch, 
                                            0.8*inch, 0.7*inch, 0.8*inch, 0.8*inch])
            conceptos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003D82')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (3, 1), (3, -1), 'LEFT'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
                ('TOPPADDING', (0, 0), (-1, 0), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
            ]))
            
            story.append(conceptos_table)
            story.append(Spacer(1, 0.15*inch))
            
            # ========== TOTALES ==========
            totales_data = [
                [Paragraph("<b>Subtotal:</b>", info_label_style), 
                Paragraph(f"${self.factura_subtotal:.2f}", info_value_style)],
                [Paragraph("<b>Descuentos:</b>", info_label_style), 
                Paragraph("$0.00", info_value_style)],
                [Paragraph("<b>Impuestos Traslados:</b>", info_label_style), 
                Paragraph(f"${self.factura_impuestos:.2f}", info_value_style)],
                [Paragraph("<b>Impuestos Retenciones:</b>", info_label_style), 
                Paragraph("$0.00", info_value_style)],
            ]
            
            totales_table = Table(totales_data, colWidths=[2.2*inch, 2.8*inch])
            totales_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F0F0F0')),
            ]))
            
            story.append(totales_table)
            story.append(Spacer(1, 0.08*inch))
            
            # ========== TOTAL PRINCIPAL ==========
            total_principal_table = Table([
                [Paragraph("<b style='font-size: 14'>TOTAL</b>", 
                        ParagraphStyle('t', parent=styles['Normal'], fontSize=14, 
                                        fontName='Helvetica-Bold')), 
                Paragraph(f"<b style='font-size: 14'>${self.factura_total:.2f}</b>", 
                        ParagraphStyle('t', parent=styles['Normal'], fontSize=14, 
                                        fontName='Helvetica-Bold'))]
            ], colWidths=[2.2*inch, 2.8*inch])
            
            total_principal_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003D82')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (0, 0), 'RIGHT'),
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            story.append(total_principal_table)
            story.append(Spacer(1, 0.2*inch))
            
            # ========== INFORMACIÓN FISCAL ==========
            story.append(Paragraph("INFORMACIÓN FISCAL", section_header_style))
            
            fiscal_data = [
                [Paragraph("<b>Serie del Certificado del Emisor:</b>", info_label_style), 
                Paragraph("30001000000400024438", info_value_style)],
                [Paragraph("<b>Folio Fiscal (UUID):</b>", info_label_style), 
                Paragraph(str(self.tfd_uuid or 'N/A'), info_value_style)],
                [Paragraph("<b>Serie del Certificado del SAT:</b>", info_label_style), 
                Paragraph("S0000001000000000028279", info_value_style)],
                [Paragraph("<b>Fecha y hora de certificación:</b>", info_label_style), 
                Paragraph(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", info_value_style)],
            ]
            
            fiscal_table = Table(fiscal_data, colWidths=[2.2*inch, 2.8*inch])
            fiscal_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F0F0F0')),
            ]))
            
            story.append(fiscal_table)
            story.append(Spacer(1, 0.1*inch))
            
            # ========== MENSAJE LEGAL ==========
            legal_text = "Este documento es una representación impresa de un CFDI"
            story.append(Paragraph(f"<i>{legal_text}</i>", 
                                ParagraphStyle('legal', parent=styles['Normal'], 
                                                fontSize=9, alignment=TA_CENTER, 
                                                textColor=colors.grey)))
            
            # ========== CONSTRUIR PDF ==========
            doc.build(story)
            
            pdf_content = buffer.getvalue()
            buffer.close()
            
            # ========== CREAR ATTACHMENT ==========
            filename = f"CFDI_{self.tfd_uuid}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            attachment = self.env['ir.attachment'].create({
                'name': filename,
                'type': 'binary',
                'datas': base64.b64encode(pdf_content),
                'res_model': self._name,
                'res_id': self.id,
                'mimetype': 'application/pdf',
            })
            
            _logger.info(f'PDF CFDI generado: {filename}')
            
            # ========== DESCARGAR AUTOMÁTICAMENTE ==========
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'self',
            }
            
        except Exception as e:
            _logger.error(f'Error generando PDF CFDI: {str(e)}')
            raise UserError(_(f'Error generando PDF CFDI: {str(e)}'))


    @api.model
    def get_available_fields(self, *args, **kwargs):
        """
        Retorna lista de campos disponibles para mostrar en la vista
        """
        fields_list = [
            {'name': 'factura_fecha', 'label': 'Fecha Comprobante', 'type': 'date'},
            {'name': 'factura_tipo', 'label': 'Tipo Comprobante', 'type': 'char'},
            {'name': 'factura_serie', 'label': 'Serie', 'type': 'char'},
            {'name': 'factura_folio', 'label': 'Folio', 'type': 'char'},
            {'name': 'rfc_emisor', 'label': 'RFC Emisor', 'type': 'char'},
            {'name': 'nombre_emisor', 'label': 'Nombre Emisor', 'type': 'char'},
            {'name': 'rfc_receptor', 'label': 'RFC Receptor', 'type': 'char'},
            {'name': 'nombre_receptor', 'label': 'Nombre Receptor', 'type': 'char'},
            {'name': 'factura_moneda', 'label': 'Moneda', 'type': 'char'},
            {'name': 'factura_tipo_cambio', 'label': 'Tipo Cambio', 'type': 'float'},
            {'name': 'factura_subtotal', 'label': 'Subtotal', 'type': 'float'},
            {'name': 'factura_impuestos', 'label': 'Impuestos', 'type': 'float'},
            {'name': 'factura_total', 'label': 'Total', 'type': 'float'},
            {'name': 'sat_status', 'label': 'Estado SAT', 'type': 'selection'},
            {'name': 'add_status', 'label': 'Estado ADD', 'type': 'selection'},
            {'name': 'document_type', 'label': 'Tipo Documento', 'type': 'selection'},
            {'name': 'document_group', 'label': 'Grupo Documento', 'type': 'selection'},
        ]

        return fields_list

    @api.model  
    def action_export_to_excel(self, invoice_ids, visible_fields):
        """
        Exporta facturas seleccionadas a Excel con campos personalizados
        """
        if not openpyxl:
            raise UserError(_("La librería openpyxl no está instalada"))
        
        invoices = self.browse(invoice_ids)
        
        if not invoices:
            raise UserError(_("No hay facturas para exportar"))
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas"
        
        # Estilos
        header_fill = PatternFill(start_color="003D82", end_color="003D82", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Obtener labels de campos
        available_fields = self.get_available_fields()
        field_labels = {f['name']: f['label'] for f in available_fields}
        
        # Escribir encabezados
        for col_idx, field_name in enumerate(visible_fields, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = field_labels.get(field_name, field_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Escribir datos
        for row_idx, invoice in enumerate(invoices, 2):
            for col_idx, field_name in enumerate(visible_fields, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                try:
                    value = invoice[field_name]
                    
                    if isinstance(value, models.BaseModel):
                        cell.value = value.display_name if value else ''
                    elif isinstance(value, datetime):
                        cell.value = value.strftime('%Y-%m-%d %H:%M:%S') if value else ''
                    elif hasattr(value, 'strftime'):  # Date objects
                        cell.value = value.strftime('%Y-%m-%d') if value else ''
                    elif isinstance(value, (int, float)):
                        cell.value = float(value) if value else 0
                        if 'total' in field_name or 'subtotal' in field_name or 'impuesto' in field_name:
                            cell.number_format = '$#,##0.00'
                    elif isinstance(value, bool):
                        cell.value = 'Sí' if value else 'No'
                    else:
                        cell.value = str(value) if value else ''
                        
                except Exception as e:
                    _logger.warning(f"Error exportando campo {field_name}: {e}")
                    cell.value = ''
                
                cell.border = border
        
        # Ajustar anchos
        for col_idx in range(1, len(visible_fields) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
        
        # Guardar
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"Facturas_{fields.Date.today().strftime('%Y%m%d')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(buffer.read()),
            'res_model': self._name,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        
        _logger.info(f'Excel exportado: {filename}')
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def get_invoice_preview_data(self):
        """
        Retorna los datos estructurados para la vista preliminar
        """
        self.ensure_one()
        
        conceptos = []
        for concepto in self.concepts_ids:
            conceptos.append({
                'id': concepto.id,
                'clave': concepto.name or '',
                'descripcion': concepto.description or '',
                'cantidad': f"{concepto.cantidad:.2f}" if concepto.cantidad else '0.00',
                'valor_unitario': f"${concepto.valor_unitario:.2f}" if concepto.valor_unitario else '$0.00',
                'importe': f"${concepto.importe:.2f}" if concepto.importe else '$0.00',
            })
        
        return {
            # Emisor
            'emisor_rfc': self.rfc_emisor or 'N/A',
            'emisor_nombre': self.nombre_emisor or 'N/A',
            
            # Receptor
            'receptor_rfc': self.rfc_receptor or 'N/A',
            'receptor_nombre': self.nombre_receptor or 'N/A',
            
            # Comprobante
            'version': self.factura_version or '4.0',
            'serie': self.factura_serie or 'N/A',
            'folio': self.factura_folio or 'N/A',
            'fecha': str(self.factura_fecha) if self.factura_fecha else 'N/A',
            'forma_pago': self.factura_forma_pago or 'N/A',
            'metodo_pago': self.factura_metodo_pago or 'N/A',
            
            # Conceptos
            'conceptos': conceptos,
            
            # Totales
            'subtotal': f"${self.factura_subtotal:.2f}",
            'impuestos': f"${self.factura_impuestos:.2f}",
            'total': f"${self.factura_total:.2f}",
            
            # Fiscal
            'uuid': self.tfd_uuid or 'N/A',
        }
    


    @api.model
    def action_generate_xml_pdf_with_fields(self, invoice_id, visible_fields):
        """
        Genera PDF respetando los campos visibles seleccionados por el usuario
        e incluye el código QR generado en tiempo real.
        """
        invoice = self.browse(invoice_id)
        if not invoice.exists():
            return False

        try:
            # --- CONFIGURACIÓN DEL CANVAS ---
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                rightMargin=0.4 * inch,
                leftMargin=0.4 * inch,
                topMargin=0.4 * inch,
                bottomMargin=0.4 * inch
            )

            styles = getSampleStyleSheet()
            story = []

            primary_color = colors.HexColor("#003D82")
            
            style_title = ParagraphStyle('Title', parent=styles['Normal'], fontSize=16, leading=20, textColor=primary_color, fontName='Helvetica-Bold')
            style_normal = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=8, leading=10)
            style_bold = ParagraphStyle('Bold', parent=styles['Normal'], fontSize=8, leading=10, fontName='Helvetica-Bold')
            style_right = ParagraphStyle('Right', parent=styles['Normal'], fontSize=8, leading=10, alignment=TA_RIGHT)
            
            # --- ENCABEZADO  ---
            if visible_fields.get('comprobante', True) or visible_fields.get('serie', True):
                header_data = [
                    [
                        Paragraph("<b>DW IT Services</b><br/><font size=7 color=grey>Software empresarial</font>", style_title),
                        Paragraph(f"""
                            <b>Tipo de Comprobante:</b> {invoice.factura_tipo or 'I'} - Ingreso<br/>
                            <b>Serie:</b> {invoice.factura_serie or ''}<br/>
                            <b>Folio:</b> {invoice.factura_folio or ''}<br/>
                            <b>Fecha:</b> {invoice.factura_fecha}<br/>
                            <b>Lugar Exp.:</b> {invoice.factura_lugar_expedicion or 'N/A'}
                        """, style_right)
                    ]
                ]
                t_header = Table(header_data, colWidths=[4*inch, 3.5*inch])
                t_header.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
                story.append(t_header)
                story.append(Spacer(1, 0.2*inch))

            # --- EMISOR Y RECEPTOR ---
            data_participants = []
            
            # Columna Emisor
            text_emisor = "<b>EMISOR:</b><br/>"
            if visible_fields.get('emisorNombre'): text_emisor += f"{invoice.nombre_emisor}<br/>"
            if visible_fields.get('emisorRFC'): text_emisor += f"<b>RFC:</b> {invoice.rfc_emisor}<br/>"
            if visible_fields.get('emisorRegimenFiscal'): text_emisor += f"Régimen Fiscal: 601"

            # Columna Receptor
            text_receptor = "<b>RECEPTOR:</b><br/>"
            if visible_fields.get('receptorNombre'): text_receptor += f"{invoice.nombre_receptor}<br/>"
            if visible_fields.get('receptorRFC'): text_receptor += f"<b>RFC:</b> {invoice.rfc_receptor}<br/>"
            if visible_fields.get('receptorUsoCFDI'): text_receptor += f"Uso CFDI: G03<br/>"
            
            data_participants.append([Paragraph(text_emisor, style_normal), Paragraph(text_receptor, style_normal)])
            
            t_participants = Table(data_participants, colWidths=[3.75*inch, 3.75*inch])
            t_participants.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F9F9F9")),
                ('PADDING', (0,0), (-1,-1), 6),
            ]))
            story.append(t_participants)
            story.append(Spacer(1, 0.2*inch))

            # --- CONCEPTOS ---
            if visible_fields.get('conceptos', True):
                headers = []
                row_widths = []
                
                if visible_fields.get('conceptoCantidad'): headers.append('Cant'); row_widths.append(0.6*inch)
                if visible_fields.get('conceptoUnidad'): headers.append('Unidad'); row_widths.append(0.8*inch)
                if visible_fields.get('conceptoClave'): headers.append('Clave'); row_widths.append(0.8*inch)
                if visible_fields.get('conceptoDescripcion'): headers.append('Descripción'); row_widths.append(2.7*inch) # Ancho variable
                if visible_fields.get('conceptoValorUnitario'): headers.append('P. Unitario'); row_widths.append(0.9*inch)
                if visible_fields.get('conceptoImporte'): headers.append('Importe'); row_widths.append(0.9*inch)
                
                total_width = sum(row_widths)
                if total_width < 7.5 * inch and 'Descripción' in headers:
                    idx_desc = headers.index('Descripción')
                    row_widths[idx_desc] += (7.5*inch - total_width)

                data_conceptos = [[Paragraph(f"<b>{h}</b>", style_bold) for h in headers]]

                for concept in invoice.concepts_ids:
                    row = []
                    if visible_fields.get('conceptoCantidad'): row.append(Paragraph(str(concept.cantidad), style_normal))
                    if visible_fields.get('conceptoUnidad'): row.append(Paragraph(str(concept.clave_unidad or ''), style_normal))
                    if visible_fields.get('conceptoClave'): row.append(Paragraph(str(concept.name or ''), style_normal))
                    if visible_fields.get('conceptoDescripcion'): row.append(Paragraph(str(concept.description or ''), style_normal))
                    if visible_fields.get('conceptoValorUnitario'): row.append(Paragraph(f"${concept.valor_unitario:,.2f}", style_right))
                    if visible_fields.get('conceptoImporte'): row.append(Paragraph(f"${concept.importe:,.2f}", style_right))
                    data_conceptos.append(row)

                t_conceptos = Table(data_conceptos, colWidths=row_widths)
                t_conceptos.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                    ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('PADDING', (0,0), (-1,-1), 4),
                ]))
                story.append(t_conceptos)
                story.append(Spacer(1, 0.1*inch))

            # --- TOTALES ---
            data_totales = []
            if visible_fields.get('subtotal'):
                data_totales.append(['Subtotal:', f"${invoice.factura_subtotal:,.2f}"])
            if visible_fields.get('impuestos'):
                data_totales.append(['Impuestos:', f"${invoice.factura_impuestos:,.2f}"])
            if visible_fields.get('total'):
                data_totales.append(['TOTAL:', f"${invoice.factura_total:,.2f}"])

            if data_totales:
                t_totales = Table(data_totales, colWidths=[6*inch, 1.5*inch])
                t_totales.setStyle(TableStyle([
                    ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
                    ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'), 
                    ('LINEABOVE', (0,-1), (-1,-1), 1, colors.black),
                ]))
                story.append(t_totales)
                story.append(Spacer(1, 0.3*inch))

            # ---  CÓDIGO QR Y SELLOS ---
            # Generar string de validación SAT
            qr_string = f"https://verificacfdi.facturaelectronica.sat.gob.mx/default.aspx?id={invoice.tfd_uuid}&re={invoice.rfc_emisor}&rr={invoice.rfc_receptor}&tt={invoice.factura_total}&fe={str(invoice.tfd_uuid)[-8:]}"
            
            # Generar imagen QR 
            qr = qrcode.QRCode(version=1, box_size=10, border=1)
            qr.add_data(qr_string)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")
            
            # Convertir imagen PIL a BytesIO 
            qr_buffer = io.BytesIO()
            qr_img.save(qr_buffer, format="PNG")
            qr_buffer.seek(0)
            

            # Datos Fiscales
            sello_info = f"""
            <b>Folio Fiscal (UUID):</b> {invoice.tfd_uuid}<br/>
            <b>No. de serie del Certificado del SAT:</b> 00001000000509846663<br/>
            <b>Fecha y hora de certificación:</b> {fields.Datetime.now()}<br/>
            <b>Sello digital del CFDI:</b><br/><font size=5>{(invoice.tfd_uuid or '') * 5}</font>
            """
            
            # Tabla Footer: [ QR | Info ]
            data_footer = [[
                Image(qr_buffer, width=1.2*inch, height=1.2*inch),
                Paragraph(sello_info, style_normal)
            ]]
            
            t_footer = Table(data_footer, colWidths=[1.5*inch, 6*inch])
            t_footer.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('ALIGN', (0,0), (0,0), 'CENTER'),
            ]))
            story.append(t_footer)

            # Leyenda
            story.append(Spacer(1, 0.1*inch))
            story.append(Paragraph("Este documento es una representación impresa de un CFDI", ParagraphStyle('Center', parent=styles['Normal'], alignment=TA_CENTER)))

            # --- CONSTRUIR ---
            doc.build(story)
            pdf_content = buffer.getvalue()
            buffer.close()

            # Guardar attachment
            filename = f"CFDI_{invoice.tfd_uuid}.pdf"
            attachment = self.env['ir.attachment'].create({
                'name': filename,
                'type': 'binary',
                'datas': base64.b64encode(pdf_content),
                'res_model': self._name,
                'res_id': self.id,
                'mimetype': 'application/pdf',
            })

            return {
                'url': f'/web/content/{attachment.id}?download=true',
                'filename': filename
            }

        except Exception as e:
            _logger.error(f"Error generando PDF: {e}")
            return False
        

    def action_view_source_document(self):
        self.ensure_one()
        # Buscar OC vinculada por UUID
        po = self.env['purchase.order'].search([('uuid', '=', self.tfd_uuid)], limit=1)
        if po:
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'purchase.order',
                'res_id': po.id,
                'view_mode': 'form',
                'target': 'current',
            }
        return {'type': 'ir.actions.client', 'tag': 'display_notification', 'params': {'message': 'No se encontró el documento origen', 'type': 'warning'}}





class SATInvoicesViewsReport(models.Model):
    _name = 'sat.invoices.views'
    _description = "SAT XML Invoices Views"
    _auto = False
    _order = 'factura_fecha desc'

    type = fields.Selection([('xmlrecibidos', 'XML Recibidos'), ('xmlemitidos', 'XML Emitidos'), ('otros', 'Otros')],
                            string='Tipo', readonly=True)
    group = fields.Selection(
        [('facturas', 'Facturas'), ('pagos', 'Pagos'), ('nominas', 'Nóminas'), ('retenciones', 'Retenciones')],
        string='Grupo', readonly=True)
    view = fields.Selection([('encabezado', 'Encabezado'), ('movimientos', 'Movimientos'), ('impuestos', 'Impuestos')],
                            string='Vista', readonly=True)
    name = fields.Char('UUID', required=True)

    factura_fecha = fields.Date('Fecha Comprobante')
    factura_tipo = fields.Char('Tipo de Comprobante')
    factura_serie = fields.Char('Serie')  # Serie
    factura_folio = fields.Char('Folio')  # Folio
    rfc_emisor = fields.Char('RFC Emisor')
    nombre_emisor = fields.Char('Nombre Emisor')
    rfc_receptor = fields.Char('RFC Receptor')
    nombre_receptor = fields.Char('Nombre Receptor')
    factura_total = fields.Float('Total')

    def init(self):
        tools.drop_view_if_exists(self._cr, 'sat_invoices_views')
        self._cr.execute('''
        DROP TABLE IF EXISTS sat_invoices_views;
        CREATE OR REPLACE VIEW sat_invoices_views AS (
            SELECT
                invoice.id,
                'xmlrecibidos' AS type,
                'facturas' AS group,
                'encabezado' AS view,
                invoice.tfd_uuid AS name,
                invoice.factura_fecha,
                invoice.factura_tipo,
                invoice.factura_serie,
                invoice.factura_folio,
                invoice.rfc_emisor,
                invoice.nombre_emisor,
                invoice.rfc_receptor,
                invoice.nombre_receptor,
                invoice.factura_total
            FROM sat_xml_invoices invoice
            UNION ALL
            SELECT
                invoice.id,
                'xmlemitidos' AS type,
                'facturas' AS group,
                'encabezado' AS view,
                invoice.tfd_uuid AS name,
                invoice.factura_fecha,
                invoice.factura_tipo,
                invoice.factura_serie,
                invoice.factura_folio,
                invoice.rfc_emisor,
                invoice.nombre_emisor,
                invoice.rfc_receptor,
                invoice.nombre_receptor,
                invoice.factura_total
            FROM sat_xml_invoices invoice
        );''')

#            sat_invoices_concepts concept
#                INNER JOIN sat_xml_invoices invoice ON invoice.id = concept.sat_invoice_id
#            WHERE concept.sat_invoice_id IS NOT NULL
